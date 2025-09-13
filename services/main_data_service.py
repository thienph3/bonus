import os
import xlrd

from typing import List
from openpyxl import load_workbook

from models.main_data_model import MainData
from repositories.main_data_repository import MainDataRepository
from utils.utils import parse_date, parse_number


class MainDataService:
    def __init__(self):
        self.repository = MainDataRepository()

    # def import_data(self, file_path, log_func=print):
    #     try:
    #         df = pd.read_excel(
    #             file_path,
    #             sheet_name='Data',
    #             skiprows=13,      # Bỏ qua 13 dòng đầu
    #             usecols='A:Q'     # Chỉ lấy cột A tới Q
    #         )

    #         if df.empty:
    #             log_func("⚠️ Sheet 'main_data' is empty.")
    #             return

    #         self.repository.delete_all()
    #         log_func("🗑️ Đã xóa dữ liệu cũ trong main_data.")

    #         df.columns = [
    #             'idx', 'document_date', 'document_number', 'description', 'corresponding_account',
    #             'increase', 'decrease', 'adjust_increase', 'adjust_decrease', 'end_amount',
    #             'seasonal_code', 'payment_period', 'customer_code', 'customer_name', 'branch', 'code', 'sales_method'
    #         ]

    #         count = 0
    #         data_list = []
    #         for index, row in df.iterrows():
    #             idx = row['idx']
    #             document_date = row['document_date']
    #             document_number = None if pd.isna(row['document_number']) else row['document_number']
    #             description = None if pd.isna(row['description']) else row['description']
    #             corresponding_account = None if pd.isna(row['corresponding_account']) else row['corresponding_account']
    #             increase = None if pd.isna(row['increase']) else row['increase']
    #             decrease = None if pd.isna(row['decrease']) else row['decrease']
    #             adjust_increase = None if pd.isna(row['adjust_increase']) else row['adjust_increase']
    #             adjust_decrease = None if pd.isna(row['adjust_decrease']) else row['adjust_decrease']
    #             end_amount = row['end_amount']
    #             seasonal_code = row['seasonal_code']
    #             payment_period = None if pd.isna(row['payment_period']) else row['payment_period']
    #             customer_code = row['customer_code']
    #             customer_name = None if pd.isna(row['customer_name']) else row['customer_name']
    #             branch = row['branch']
    #             code = None if pd.isna(row['code']) else row['code']
    #             sales_method = row['sales_method']

    #             try:
    #                 idx = parse_number(idx)
    #                 document_date = parse_date(document_date)
    #                 increase = parse_number(increase)
    #                 decrease = parse_number(decrease)
    #                 adjust_increase = parse_number(adjust_increase)
    #                 adjust_decrease = parse_number(adjust_decrease)
    #                 end_amount = parse_number(end_amount)
    #                 payment_period = parse_number(payment_period)

    #                 data = {
    #                     "idx": idx,
    #                     "document_date": document_date,
    #                     "document_number": document_number,
    #                     "description": description,
    #                     "corresponding_account": corresponding_account,
    #                     "increase": increase,
    #                     "decrease": decrease,
    #                     "adjust_increase": adjust_increase,
    #                     "adjust_decrease": adjust_decrease,
    #                     "end_amount": end_amount,
    #                     "seasonal_code": seasonal_code,
    #                     "payment_period": payment_period,
    #                     "customer_code": customer_code,
    #                     "customer_name": customer_name,
    #                     "branch": branch,
    #                     "code": code,
    #                     "sales_method": sales_method
    #                 }
    #                 data_list.append(data)
    #                 if len(data_list) % 100 == 0:
    #                     error_indexes = self.repository.bulk_create(data_list, 0, 100)
    #                     count += len(data_list) - len(error_indexes)
    #                     log_func(f"Đã nhập first {count}/{len(df)} rows.")
    #                     for i in error_indexes:
    #                         log_func(f"❌ Exception in row: {data_list[i]}")
    #                     data_list = []

    #             except ValueError as ve:
    #                 log_func(f"⚠️ ValueError in row {index + 1}: {ve}. Skipping. Row error: {row.to_dict()}")
    #                 continue
    #             except Exception as e:
    #                 log_func(f"❌ Exception in row {index + 1}: {e}. Skipping. Row error: {row.to_dict()}")
    #                 continue

    #         log_func(f"✅ Đã nhập {count}/{len(df)} dòng dữ liệu vào main_data.")

    #     except Exception as e:
    #         log_func(f"❌ An error occurred during import: {e}")
    #         raise
    
    def read_excel_data_dynamic(self, file_path, sheet_name="Data", max_scan_rows=30, min_non_empty=17, max_col=17):
        ext = os.path.splitext(file_path)[-1].lower()

        if ext == ".xlsx":
            wb = load_workbook(filename=file_path, read_only=True, data_only=True)
            sheet = wb[sheet_name]

            data_start_row = None
            for i, row in enumerate(sheet.iter_rows(min_row=1, max_row=max_scan_rows), start=1):
                values = [cell.value for cell in row]
                if sum(1 for v in values if v not in (None, "", " ")) >= min_non_empty:
                    data_start_row = i
                    break

            if data_start_row is None:
                raise Exception("No data found in .xlsx file")
            
            data_start_row += 1

            # Read data from data_start_row onwards
            rows = sheet.iter_rows(min_row=data_start_row, values_only=True)
            result = []
            for row in rows:
                if max_col:
                    result.append(row[:max_col])
                else:
                    result.append(row)
            return data_start_row, result

        elif ext == ".xls":
            wb = xlrd.open_workbook(file_path)
            sheet = wb.sheet_by_name(sheet_name)

            data_start_row = None
            for row_idx in range(min(sheet.nrows, max_scan_rows)):
                row = sheet.row_values(row_idx)
                if sum(1 for cell in row if str(cell).strip()) >= min_non_empty:
                    data_start_row = row_idx
                    break

            if data_start_row is None:
                raise Exception("No data found in .xls file")
            
            data_start_row += 2

            result = []
            for i in range(data_start_row, sheet.nrows):
                row = sheet.row_values(i)
                if max_col:
                    result.append(row[:max_col])
                else:
                    result.append(row)
            return data_start_row, result

        else:
            raise ValueError("Unsupported file extension: must be .xls or .xlsx")


    def import_data(self, file_path, log_func=print, batch_size=100):
        try:
            # wb = load_workbook(filename=file_path, read_only=True, data_only=True)
            # sheet = wb["Data"]

            # # Skip first 13 rows (header)
            # rows = sheet.iter_rows(min_row=15, max_col=17, values_only=True)
            data_start_row, rows = self.read_excel_data_dynamic(file_path)

            self.repository.delete_all()
            log_func("🗑️ Đã xóa dữ liệu cũ trong main_data.")

            headers = [
                "idx",
                "document_date",
                "document_number",
                "description",
                "corresponding_account",
                "increase",
                "decrease",
                "adjust_increase",
                "adjust_decrease",
                "end_amount",
                "seasonal_code",
                "payment_period",
                "customer_code",
                "customer_name",
                "branch",
                "code",
                "sales_method",
            ]

            batch = []
            total = 0
            for excel_index, row in enumerate(
                rows, start=data_start_row
            ):  # Excel index starts at 14 (row number)
                if all(
                    cell is None or (isinstance(cell, str) and cell.strip() == "")
                    for cell in row
                ):
                    break

                try:
                    row_dict = dict(zip(headers, row))
                    data = {
                        "idx": parse_number(row_dict["idx"]),
                        "document_date": parse_date(row_dict["document_date"]),
                        "document_number": row_dict["document_number"],
                        "description": row_dict["description"],
                        "corresponding_account": row_dict["corresponding_account"],
                        "increase": parse_number(row_dict["increase"]),
                        "decrease": parse_number(row_dict["decrease"]),
                        "adjust_increase": parse_number(row_dict["adjust_increase"]),
                        "adjust_decrease": parse_number(row_dict["adjust_decrease"]),
                        "end_amount": parse_number(row_dict["end_amount"]),
                        "seasonal_code": row_dict["seasonal_code"],
                        "payment_period": parse_number(row_dict["payment_period"]),
                        "customer_code": row_dict["customer_code"],
                        "customer_name": row_dict["customer_name"],
                        "branch": row_dict["branch"],
                        "code": row_dict["code"],
                        "sales_method": row_dict["sales_method"],
                    }
                    batch.append((excel_index, data))

                    if len(batch) >= batch_size:
                        batch_data = [d for (_, d) in batch]
                        error_indexes = self.repository.bulk_create(batch_data)
                        total += len(batch_data) - len(error_indexes)
                        log_func(f"✅ Đã nhập {total} dòng dữ liệu vào main-data...")

                        for i in error_indexes:
                            row_num, error_row = batch[i]
                            log_func(f"❌ Insert failed at row {row_num}: {error_row}")

                        batch.clear()

                except Exception as e:
                    log_func(f"⚠️ Parse error at row {excel_index}: {e}")
                    self.repository.rollback()
                    continue

            # Insert remaining rows
            if batch:
                batch_data = [d for (_, d) in batch]
                error_indexes = self.repository.bulk_create(batch_data)
                total += len(batch_data) - len(error_indexes)
                log_func(f"✅ Đã nhập {total} dòng dữ liệu vào main-data...")
                for i in error_indexes:
                    row_num, error_row = batch[i]
                    log_func(f"❌ Insert failed at row {row_num}: {error_row}")

        except Exception as e:
            log_func(f"❌ Fatal error: {e}")
            raise

    def get_all(self) -> List[MainData]:
        return self.repository.get_all()

    def get_page(self, page, page_size, filters, sort):
        items, total = self.repository.get_page(page, page_size)
        return [item.to_dict() for item in items], total

    def get(self, id):
        item = self.repository.get_by_id(id)
        return item.to_dict()

    # def add(self, name, date, desc):
    #     return self.repository.create(name, date, desc)

    def update(self, id, **kwargs):
        return self.repository.update(id, **kwargs)

    def create(self, **kwargs):
        return self.repository.create(**kwargs)

    def remove(self, id):
        return self.repository.delete(id)

    def rollback(self):
        return self.repository.rollback()
